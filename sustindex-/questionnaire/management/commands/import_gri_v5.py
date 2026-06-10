# -*- coding: utf-8 -*-
"""
Import GRI v5 structured Excel (Bolum 1-4 + Sector) into the database.

Usage
-----
  # Import everything in the file:
  python manage.py import_gri_v5 "C:/path/to/GRI_v5_Bolum1_Governance.xlsx"

  # Import only specific criteria (end-to-end test):
  python manage.py import_gri_v5 "C:/path/to/file.xlsx" --criteria G1
  python manage.py import_gri_v5 "C:/path/to/file.xlsx" --criteria G1 G2 G3

  # Wipe matching surveys first:
  python manage.py import_gri_v5 "C:/path/to/file.xlsx" --clear

Excel column layout (row 1 = header):
  A  survey_name
  B  criterion_code          e.g. G1
  C  criterion_title         e.g. Board Oversight & Governance
  D  gri_reference           e.g. GRI 2-9, 2-10
  E  mandatory               Yes / No
  F  layer                   GATE | P | I | M | R | CONDITIONAL
  G  order_in_criterion      1, 2, 3 ...
  H  question_type           binary | single | multi | numerical | text
  I  is_gate                 Yes / No
  J  is_conditional          Yes / No
  K  conditional_on_code     e.g. G1-GATE  (criterion_code + "-" + layer of parent)
  L  conditional_on_min_score integer (default 1)
  M  bonus_points            integer (default 0)
  N  layer_max_pts           integer (0 for GATE / CONDITIONAL rows)
  O  question_text_tr        Turkish
  P  question_text_en        English
  Q  numerical_thresholds_json  JSON string or blank
  R  opt1_text  S opt1_score
  T  opt2_text  U opt2_score
  V  opt3_text  W opt3_score
  X  opt4_text  Y opt4_score
  Z  opt5_text  AA opt5_score
  AB opt6_text  AC opt6_score
  AD notes
"""

import json
import re
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from questionnaire.models import Survey, Category, Question, Choice


# ── helpers ────────────────────────────────────────────────────────────────

def _s(v):
    if v is None:
        return ''
    return str(v).strip()


def _n(v, default=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _bool(v):
    return _s(v).lower() in ('yes', '1', 'true', 'evet')


# Column indices (0-based, matching the header layout above)
COL = {
    'survey_name':              0,
    'criterion_code':           1,
    'criterion_title':          2,
    'gri_reference':            3,
    'mandatory':                4,
    'layer':                    5,
    'order_in_criterion':       6,
    'question_type':            7,
    'is_gate':                  8,
    'is_conditional':           9,
    'conditional_on_code':      10,
    'conditional_on_min_score': 11,
    'bonus_points':             12,
    'layer_max_pts':            13,
    'question_text_tr':         14,
    'question_text_en':         15,
    'numerical_thresholds_json':16,
    'opt1_text':  17, 'opt1_score': 18,
    'opt2_text':  19, 'opt2_score': 20,
    'opt3_text':  21, 'opt3_score': 22,
    'opt4_text':  23, 'opt4_score': 24,
    'opt5_text':  25, 'opt5_score': 26,
    'opt6_text':  27, 'opt6_score': 28,
    'notes':      29,
}


def _safe(text):
    """Return ASCII-safe version of text for stdout (Windows cp1252 safe)."""
    return text.encode('ascii', errors='replace').decode('ascii')


def _get(row, key, default=''):
    idx = COL[key]
    val = row[idx] if idx < len(row) else None
    return val if val is not None else default


class Command(BaseCommand):
    help = 'Import GRI v5 structured Excel (Bolum 1-4 + Sector).'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str)
        parser.add_argument(
            '--criteria', nargs='+', default=[],
            help='Import only these criterion codes, e.g. --criteria G1 G2',
        )
        parser.add_argument(
            '--clear', action='store_true',
            help='Delete matching surveys before import (matched by survey_name in Excel)',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl not installed. Run: pip install openpyxl')

        path = options['excel_path']
        filter_criteria = set(c.upper() for c in options['criteria'])
        self.stdout.write(f'[v5] Loading: {path}')

        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {path}')
        except Exception as e:
            raise CommandError(f'Cannot open file: {e}')

        try:
            self._run(wb, filter_criteria, options['clear'])
        finally:
            wb.close()

    def _run(self, wb, filter_criteria, do_clear):
        # ── 1. Parse all sheets into a flat list of row-dicts ────────────────
        all_rows = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                if _s(row[0]).lower() == 'survey_name':
                    header_row = row
                    break
            if header_row is None:
                self.stdout.write(self.style.WARNING(
                    f'  [!] Sheet "{sheet_name}" has no header row - skipped.'
                ))
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                r = list(row) + [None] * 35          # pad to avoid index errors
                survey_name = _s(r[COL['survey_name']])
                criterion   = _s(r[COL['criterion_code']]).upper()
                if not survey_name or not criterion:
                    continue
                if filter_criteria and criterion not in filter_criteria:
                    continue
                all_rows.append(r)

        if not all_rows:
            self.stdout.write(self.style.WARNING('[!] No data rows matched - nothing imported.'))
            return

        # ── 2. Group by survey → criterion ───────────────────────────────────
        from collections import defaultdict, OrderedDict
        surveys_data = OrderedDict()   # survey_name → {criterion_code → [rows]}
        for r in all_rows:
            sn = _s(r[COL['survey_name']])
            cc = _s(r[COL['criterion_code']]).upper()
            surveys_data.setdefault(sn, OrderedDict())
            surveys_data[sn].setdefault(cc, [])
            surveys_data[sn][cc].append(r)

        # ── 3. Optionally clear ──────────────────────────────────────────────
        if do_clear:
            for sn in surveys_data:
                n, _ = Survey.objects.filter(name=sn).delete()
                if n:
                    self.stdout.write(self.style.WARNING(f'  [!] Deleted survey: {sn}'))

        # ── 4. Upsert surveys, categories, questions, choices ────────────────
        total_q = total_c = 0

        # We need a two-pass approach per survey to resolve conditional FK links.
        # Pass 1: create/update all questions (without conditional FK)
        # Pass 2: wire up conditional_on_question FK

        for survey_name, criteria_dict in surveys_data.items():
            survey = self._upsert_survey(survey_name)

            # Maps: (criterion_code, layer) → Question object
            q_by_code_layer = {}

            # Pass 1 ─ questions + choices
            cat_order = 0
            for criterion_code, rows in criteria_dict.items():
                cat_order += 1
                # Derive category metadata from first row
                first = rows[0]
                cat_title = _s(first[COL['criterion_title']]) or criterion_code
                gri_ref   = _s(first[COL['gri_reference']])
                mandatory = _bool(first[COL['mandatory']])

                cat = self._upsert_category(
                    survey, criterion_code, cat_title, gri_ref, mandatory, cat_order
                )

                for r in rows:
                    layer         = _s(r[COL['layer']])
                    order         = _n(r[COL['order_in_criterion']], default=99)
                    qtype         = _s(r[COL['question_type']]) or 'single'
                    is_gate       = _bool(r[COL['is_gate']])
                    layer_max     = _n(r[COL['layer_max_pts']])
                    text_tr       = _s(r[COL['question_text_tr']])
                    text_en       = _s(r[COL['question_text_en']])
                    bonus_pts     = _n(r[COL['bonus_points']])
                    cond_min      = _n(r[COL['conditional_on_min_score']], default=1)
                    thresholds_js = _s(r[COL['numerical_thresholds_json']])
                    allow_multi   = (qtype == 'multi')

                    # Parse numerical thresholds
                    thresholds = None
                    if thresholds_js:
                        try:
                            thresholds = json.loads(thresholds_js)
                        except json.JSONDecodeError:
                            self.stdout.write(self.style.WARNING(
                                f'  [!] Bad JSON thresholds for {criterion_code} layer={layer}'
                            ))

                    # Normalise legacy question types
                    if qtype == 'choice':
                        qtype = 'single'

                    q_obj, new_q = Question.objects.update_or_create(
                        survey=survey,
                        category=cat,
                        order=order,
                        defaults={
                            'text':                  text_tr or text_en,
                            'text_tr':               text_tr,
                            'text_en':               text_en,
                            'question_type':         qtype,
                            'allow_multiple':        allow_multi,
                            'is_active':             True,
                            # new v5 fields
                            'criterion_code':        criterion_code,
                            'layer':                 layer,
                            'is_gate':               is_gate,
                            'numerical_thresholds':  thresholds,
                            'bonus_points':          bonus_pts,
                            'conditional_on_min_score': cond_min,
                            # conditional_on_question wired in pass 2
                        }
                    )

                    if new_q:
                        total_q += 1

                    # Track for cross-referencing
                    q_by_code_layer[(criterion_code, layer)] = q_obj

                    # Upsert choices (binary / single / multi)
                    if qtype not in ('numerical', 'text'):
                        for slot in range(1, 7):
                            opt_text  = _s(r[COL[f'opt{slot}_text']])
                            opt_score = _n(r[COL[f'opt{slot}_score']])
                            if not opt_text:
                                continue
                            _, new_c = Choice.objects.update_or_create(
                                question=q_obj,
                                order=slot,
                                defaults={
                                    'text':    opt_text,
                                    'text_en': opt_text,
                                    'text_tr': opt_text,
                                    'score':   opt_score,
                                }
                            )
                            if new_c:
                                total_c += 1

            # Pass 2 ─ wire conditional FK
            for r in [row for rows in criteria_dict.values() for row in rows]:
                cond_on_code = _s(r[COL['conditional_on_code']])
                if not cond_on_code:
                    continue

                # Parse "G1-GATE" or "G1-P" format
                parts = cond_on_code.split('-', 1)
                if len(parts) != 2:
                    continue
                parent_criterion, parent_layer = parts[0].upper(), parts[1].upper()
                parent_q = q_by_code_layer.get((parent_criterion, parent_layer))
                if parent_q is None:
                    self.stdout.write(self.style.WARNING(
                        f'  [!] conditional_on_code "{cond_on_code}" not found - skipping link'
                    ))
                    continue

                # Find the child question
                criterion_code = _s(r[COL['criterion_code']]).upper()
                layer          = _s(r[COL['layer']])
                child_q = q_by_code_layer.get((criterion_code, layer))
                if child_q and child_q.conditional_on_question_id != parent_q.pk:
                    Question.objects.filter(pk=child_q.pk).update(
                        conditional_on_question=parent_q
                    )

            self.stdout.write(
                f'  [OK] Survey "{_safe(survey_name)}" -- {len(criteria_dict)} criteria'
            )

        self.stdout.write(self.style.SUCCESS(
            f'\n[DONE] {total_q} new questions, {total_c} new choices imported.\n'
        ))

    # ── upsert helpers ─────────────────────────────────────────────────────

    def _upsert_survey(self, name):
        survey, created = Survey.objects.update_or_create(
            name=name,
            defaults={
                'name_en':                  name,
                'name_tr':                  name,
                'description':              name,
                'description_en':           name,
                'description_tr':           name,
                'is_active':                True,
                'allow_multiple_attempts':  True,
                'show_results_immediately': True,
            }
        )
        return survey

    def _upsert_category(self, survey, criterion_code, title, gri_ref,
                         mandatory, order):
        full_name = f'{criterion_code}: {title}'
        desc = f'{gri_ref}' if gri_ref else ''
        cat, _ = Category.objects.update_or_create(
            survey=survey,
            name=full_name,
            defaults={
                'name_en':              full_name,
                'name_tr':              full_name,
                'order':                order,
                'max_score':            0,    # recalculated after all questions saved
                'environmental_weight': 0.0,
                'social_weight':        0.0,
                'governance_weight':    1.0,
            }
        )
        return cat
