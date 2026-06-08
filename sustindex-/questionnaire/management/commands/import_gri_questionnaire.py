"""
Import GRI Standards Assessment v4 (Structured) from Excel into the database.

Usage
-----
  python manage.py import_gri_questionnaire "C:/path/to/GRI_Questionnaire_v4_STRUCTURED.xlsx"
  python manage.py import_gri_questionnaire "C:/path/to/file.xlsx" --clear

What it creates
---------------
  11 separate Surveys following the actual GRI Universal Standards hierarchy:

  Core (phased — all companies complete all three):
    - GRI 1: Foundation                   (32 questions,  160 pts max)
    - GRI 2: General Disclosures          (80 questions,  400 pts max)
    - GRI 3: Material Topics              (60 questions,  300 pts max)

  Sector Standards (company picks ONE):
    - GRI Sector: Agriculture & Food      (8 questions)
    - GRI Sector: Energy & Utilities      (8 questions)
    - GRI Sector: Financial Services      (8 questions)
    - GRI Sector: Manufacturing & Industry(8 questions)
    - GRI Sector: Construction & Real Estate (8 questions)
    - GRI Sector: Healthcare & Pharma     (8 questions)
    - GRI Sector: Technology & IT         (8 questions)
    - GRI Sector: Retail & Trade          (8 questions)

The command is fully idempotent: re-running updates existing records.
"""

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from questionnaire.models import Survey, Category, Question, Choice


# ── Section config ─────────────────────────────────────────────────────────
# Follows the actual GRI Universal Standards hierarchy:
#   GRI 1 Foundation → GRI 2 General Disclosures → GRI 3 Material Topics
# Phased completion: every company works through all three core phases in order
# before selecting their sector-specific standard.

CORE_SECTIONS = [
    {
        'sheet':         'GRI 1 — Foundation',
        'survey_name':   'GRI 1: Foundation',
        'survey_name_tr':'GRI 1: Temel',
        'survey_desc':   'GRI 1:2021 Foundation — 8 criteria \xd7 4 layers (Policy/Implementation/Measurement/Results) = 32 questions, 160 pts max.',
        'cat_name':      'Foundation',
        'cat_name_tr':   'Temel',
        'max_score':     160,
        'env_w': 0.10, 'soc_w': 0.10, 'gov_w': 0.80,
    },
    {
        'sheet':         'GRI 2 — General Disclosures',
        'survey_name':   'GRI 2: General Disclosures',
        'survey_name_tr':'GRI 2: Genel A\xe7ıklamalar',
        'survey_desc':   'GRI 2:2021 General Disclosures — 20 criteria \xd7 4 layers = 80 questions, 400 pts max.',
        'cat_name':      'General Disclosures',
        'cat_name_tr':   'Genel A\xe7ıklamalar',
        'max_score':     400,
        'env_w': 0.15, 'soc_w': 0.15, 'gov_w': 0.70,
    },
    {
        'sheet':         'GRI 3 — Material Topics',
        'survey_name':   'GRI 3: Material Topics',
        'survey_name_tr':'GRI 3: \xd6nemli Konular',
        'survey_desc':   'GRI 3:2021 Material Topics — 15 criteria \xd7 4 layers = 60 questions, 300 pts max.',
        'cat_name':      'Material Topics',
        'cat_name_tr':   '\xd6nemli Konular',
        'max_score':     300,
        'env_w': 0.33, 'soc_w': 0.34, 'gov_w': 0.33,
    },
]

SECTOR_SECTIONS = [
    {
        'sheet':         'Sector — Agriculture & Food',
        'survey_name':   'GRI Sector: Agriculture & Food',
        'survey_name_tr':'GRI Sektor: Tarim & Gida',
        'survey_desc':   'GRI 13 — 8 sector-specific questions.',
        'cat_name':      'Agriculture & Food',
        'cat_name_tr':   'Tarim & Gida',
    },
    {
        'sheet':         'Sector — Energy & Utilities',
        'survey_name':   'GRI Sector: Energy & Utilities',
        'survey_name_tr':'GRI Sektor: Enerji & Hizmetler',
        'survey_desc':   'GRI 11 — 8 sector-specific questions.',
        'cat_name':      'Energy & Utilities',
        'cat_name_tr':   'Enerji & Hizmetler',
    },
    {
        'sheet':         'Sector — Financial Services',
        'survey_name':   'GRI Sector: Financial Services',
        'survey_name_tr':'GRI Sektor: Finansal Hizmetler',
        'survey_desc':   'PCAF / TCFD / GRI 14 — 8 sector-specific questions.',
        'cat_name':      'Financial Services',
        'cat_name_tr':   'Finansal Hizmetler',
    },
    {
        'sheet':         'Sector — Manufacturing & Indust',
        'survey_name':   'GRI Sector: Manufacturing & Industry',
        'survey_name_tr':'GRI Sektor: Imalat & Sanayi',
        'survey_desc':   'GRI 300 series — 8 sector-specific questions.',
        'cat_name':      'Manufacturing & Industry',
        'cat_name_tr':   'Imalat & Sanayi',
    },
    {
        'sheet':         'Sector — Construction & Real Es',
        'survey_name':   'GRI Sector: Construction & Real Estate',
        'survey_name_tr':'GRI Sektor: Insaat & Gayrimenkul',
        'survey_desc':   'GRESB / GRI 300 — 8 sector-specific questions.',
        'cat_name':      'Construction & Real Estate',
        'cat_name_tr':   'Insaat & Gayrimenkul',
    },
    {
        'sheet':         'Sector — Healthcare & Pharma',
        'survey_name':   'GRI Sector: Healthcare & Pharma',
        'survey_name_tr':'GRI Sektor: Saglik & Ilac',
        'survey_desc':   'IFPMA / GRI 400 — 8 sector-specific questions.',
        'cat_name':      'Healthcare & Pharma',
        'cat_name_tr':   'Saglik & Ilac',
    },
    {
        'sheet':         'Sector — Technology & IT',
        'survey_name':   'GRI Sector: Technology & IT',
        'survey_name_tr':'GRI Sektor: Teknoloji & BT',
        'survey_desc':   'GRI 418 / TCFD — 8 sector-specific questions.',
        'cat_name':      'Technology & IT',
        'cat_name_tr':   'Teknoloji & BT',
    },
    {
        'sheet':         'Sector — Retail & Trade',
        'survey_name':   'GRI Sector: Retail & Trade',
        'survey_name_tr':'GRI Sektor: Perakende & Ticaret',
        'survey_desc':   'GRI 300/400 series — 8 sector-specific questions.',
        'cat_name':      'Retail & Trade',
        'cat_name_tr':   'Perakende & Ticaret',
    },
]

LETTER_ORDER = {'A': 1, 'B': 2, 'C': 3, 'D': 4}

# Words that appear in v4 STRUCTURED Excel's "Layer" column (p[3]).
# These are NOT question texts — the parser must skip them so that
# SECTOR_QUESTION_TEXT provides the real question sentences instead.
LAYER_MARKERS = frozenset({
    'sector', 'universal', 'policy', 'implementation', 'measurement', 'results',
})

# ── Full question text for every sector question ────────────────────────────
# Covers BOTH v3 Q_IDs (AG-01, RET-01 …) and v4 Q_IDs (AGRI-1, RET-1 …).
# The v4 STRUCTURED file uses a "Layer" column (value = "Sector") in the
# position that v3 used for the full question text; LAYER_MARKERS above
# prevents that from being stored.
#
# Priority order used by _parse_sector:
#   1. Excel col-D text (if filled in AND not a layer marker)
#   2. This mapping (SECTOR_QUESTION_TEXT, keyed by normalised Q_ID)
#   3. "Category  (GRI Ref)" fallback
#
# Turkish translations are stored in text_tr and served by the frontend
# when lang='tr'.
SECTOR_QUESTION_TEXT = {
    # ════════════════════════════════════════════════════════════════════════
    # v4 STRUCTURED IDs  (AGRI-1 … RET-8)
    # ════════════════════════════════════════════════════════════════════════

    # ── Agriculture & Food (v4) ────────────────────────────────────────────
    'AGRI-1': {
        'en': "Does your organisation have a formal policy to respect and protect the land rights and tenure of local communities and indigenous peoples?",
        'tr': "Kuruluşunuzun yerel toplulukların ve yerli halkların arazi hakları ile tapu güvencesini saygıyla korumaya yönelik resmi bir politikası var mı?",
    },
    'AGRI-2': {
        'en': "How does your organisation support the inclusion of smallholder farmers and local communities in its agricultural supply chain?",
        'tr': "Kuruluşunuz küçük çiftçilerin ve yerel toplulukların tarımsal tedarik zincirine dahil edilmesini nasıl desteklemektedir?",
    },
    'AGRI-3': {
        'en': "Does your organisation measure, manage and reduce water use across its agricultural operations and supply chain?",
        'tr': "Kuruluşunuz tarımsal faaliyetleri ve tedarik zinciri genelinde su kullanımını ölçüyor, yönetiyor ve azaltıyor mu?",
    },
    'AGRI-4': {
        'en': "Does your organisation assess and mitigate its impact on biodiversity and ecosystem services in agricultural areas?",
        'tr': "Kuruluşunuz tarım alanlarında biyoçeşitlilik ve ekosistem hizmetleri üzerindeki etkisini değerlendiriyor ve azaltıyor mu?",
    },
    'AGRI-5': {
        'en': "Does your organisation have a policy for responsible pesticide and chemical management that goes beyond minimum regulatory requirements?",
        'tr': "Kuruluşunuzun asgari yasal gerekliliklerin ötesine geçen sorumlu pestisit ve kimyasal yönetimine ilişkin bir politikası var mı?",
    },
    'AGRI-6': {
        'en': "Does your organisation have a documented food safety and quality management system with third-party certification?",
        'tr': "Kuruluşunuzun üçüncü taraf sertifikasına sahip, belgelenmiş bir gıda güvenliği ve kalite yönetim sistemi var mı?",
    },
    'AGRI-7': {
        'en': "Does your organisation have measures to identify and prevent child labour and forced labour throughout its agricultural supply chain?",
        'tr': "Kuruluşunuz tarımsal tedarik zinciri boyunca çocuk işçiliğini ve zorla çalıştırmayı tespit etmeye ve önlemeye yönelik tedbirler alıyor mu?",
    },
    'AGRI-8': {
        'en': "Does your organisation measure and publicly report greenhouse gas emissions from its agricultural operations and supply chain?",
        'tr': "Kuruluşunuz tarımsal faaliyetlerinden ve tedarik zincirinden kaynaklanan sera gazı emisyonlarını ölçüp kamuoyuyla paylaşıyor mu?",
    },

    # ── Energy & Utilities (v4) ────────────────────────────────────────────
    'ENER-1': {
        'en': "Does your organisation measure, set reduction targets for and publicly report GHG emissions from energy generation?",
        'tr': "Kuruluşunuz enerji üretiminden kaynaklanan sera gazı emisyonlarını ölçüyor, azaltma hedefleri koyuyor ve kamuoyuyla paylaşıyor mu?",
    },
    'ENER-2': {
        'en': "Does your organisation have a strategy and published targets for transitioning to renewable energy sources?",
        'tr': "Kuruluşunuzun yenilenebilir enerji kaynaklarına geçiş için bir stratejisi ve yayımlanmış hedefleri var mı?",
    },
    'ENER-3': {
        'en': "Does your organisation monitor, manage and actively reduce water consumption in its energy generation processes?",
        'tr': "Kuruluşunuz enerji üretim süreçlerindeki su tüketimini izliyor, yönetiyor ve aktif olarak azaltıyor mu?",
    },
    'ENER-4': {
        'en': "Does your organisation assess and manage physical climate risks to its infrastructure, assets and operations?",
        'tr': "Kuruluşunuz altyapısı, varlıkları ve faaliyetleri üzerindeki fiziksel iklim risklerini değerlendiriyor ve yönetiyor mu?",
    },
    'ENER-5': {
        'en': "Does your organisation have a published just transition plan that protects workers and communities affected by the shift to clean energy?",
        'tr': "Kuruluşunuzun temiz enerjiye geçiş sürecinden etkilenen işçileri ve toplulukları koruyan yayımlanmış bir adil geçiş planı var mı?",
    },
    'ENER-6': {
        'en': "Does your organisation have programmes and targets to promote affordable and reliable energy access for all, including underserved populations?",
        'tr': "Kuruluşunuz, hizmet götürülemeyen kesimler dahil herkes için uygun fiyatlı ve güvenilir enerji erişimini teşvik eden programlara ve hedeflere sahip mi?",
    },
    'ENER-7': {
        'en': "Does your organisation measure, disclose and actively reduce air quality pollutant emissions (SO₂, NOₓ, PM) in line with Best Available Technology standards?",
        'tr': "Kuruluşunuz En İyi Mevcut Teknoloji standartlarına uygun biçimde hava kalitesi kirletici emisyonlarını (SO₂, NOₓ, PM) ölçüyor, ifşa ediyor ve azaltıyor mu?",
    },
    'ENER-8': {
        'en': "Does your organisation have a comprehensive nuclear safety and radioactive waste management policy and programme?",
        'tr': "Kuruluşunuzun kapsamlı bir nükleer güvenlik ve radyoaktif atık yönetimi politikası ile programı var mı?",
    },

    # ── Financial Services (v4) ────────────────────────────────────────────
    'FIN-1': {
        'en': "Does your organisation measure and disclose financed emissions using the PCAF methodology or equivalent, with third-party verification?",
        'tr': "Kuruluşunuz üçüncü taraf doğrulamasıyla PCAF metodolojisi veya eşdeğerini kullanarak finanse edilen emisyonları ölçüyor ve açıklıyor mu?",
    },
    'FIN-2': {
        'en': "Does your organisation conduct climate scenario analysis and assess climate-related financial risk across its credit and investment portfolio?",
        'tr': "Kuruluşunuz iklim senaryo analizi yapıyor ve kredi ile yatırım portföyü genelinde iklimle ilgili finansal riski değerlendiriyor mu?",
    },
    'FIN-3': {
        'en': "Does your organisation have a published responsible investment policy that integrates ESG criteria into investment and lending decisions?",
        'tr': "Kuruluşunuzun yatırım ve kredi kararlarına ESG kriterlerini entegre eden yayımlanmış bir sorumlu yatırım politikası var mı?",
    },
    'FIN-4': {
        'en': "Does your organisation have programmes and targets to promote financial inclusion and access for underserved populations?",
        'tr': "Kuruluşunuz hizmet götürülemeyen kesimlere yönelik finansal kapsayıcılığı ve erişimi teşvik eden program ve hedeflere sahip mi?",
    },
    'FIN-5': {
        'en': "Does your organisation have robust anti-money laundering and financial crime compliance systems with regular independent audits?",
        'tr': "Kuruluşunuzun düzenli bağımsız denetimlerle desteklenen sağlam kara para aklamayla mücadele ve mali suç uyum sistemleri var mı?",
    },
    'FIN-6': {
        'en': "Does your organisation have a comprehensive data privacy and cybersecurity risk management framework with published breach statistics?",
        'tr': "Kuruluşunuzun yayımlanmış ihlal istatistikleriyle desteklenen kapsamlı bir veri gizliliği ve siber güvenlik risk yönetimi çerçevesi var mı?",
    },
    'FIN-7': {
        'en': "Does your organisation disclose diversity metrics across leadership and the workforce, and have targets to improve representation?",
        'tr': "Kuruluşunuz liderlik ve işgücü genelinde çeşitlilik ölçütlerini açıklıyor ve temsili artırmaya yönelik hedeflere sahip mi?",
    },
    'FIN-8': {
        'en': "Does your organisation integrate ESG considerations into its product and service design, and offer sustainable finance products to clients?",
        'tr': "Kuruluşunuz ürün ve hizmet tasarımına ESG değerlendirmelerini dahil ediyor ve müşterilere sürdürülebilir finans ürünleri sunuyor mu?",
    },

    # ── Manufacturing & Industry (v4) ──────────────────────────────────────
    'MFG-1': {
        'en': "Does your organisation have targets and programmes to improve energy efficiency in manufacturing operations, with published performance data?",
        'tr': "Kuruluşunuz imalat operasyonlarında enerji verimliliğini artırmaya yönelik yayımlanmış performans verileriyle desteklenen hedef ve programlara sahip mi?",
    },
    'MFG-2': {
        'en': "Does your organisation measure, reduce and report waste generation, and apply circular economy principles to minimise landfill?",
        'tr': "Kuruluşunuz atık üretimini ölçüyor, azaltıyor ve raporluyor; düzenli depolama alanını en aza indirmek için döngüsel ekonomi ilkelerini uyguluyor mu?",
    },
    'MFG-3': {
        'en': "Does your organisation have a management system for the safe handling, storage and disposal of chemicals and hazardous materials?",
        'tr': "Kuruluşunuzun kimyasallar ve tehlikeli maddelerin güvenli taşınması, depolanması ve bertarafı için bir yönetim sistemi var mı?",
    },
    'MFG-4': {
        'en': "Does your organisation have a certified occupational health and safety management system with measurable targets and published incident data?",
        'tr': "Kuruluşunuzun ölçülebilir hedefler ve yayımlanmış kaza verileriyle desteklenen sertifikalı bir iş sağlığı ve güvenliği yönetim sistemi var mı?",
    },
    'MFG-5': {
        'en': "Does your organisation audit and monitor labour standards throughout its supplier base, including Tier 2 and beyond?",
        'tr': "Kuruluşunuz 2. Kademe ve ötesini kapsayan tedarikçi tabanı genelinde çalışma standartlarını denetliyor ve izliyor mu?",
    },
    'MFG-6': {
        'en': "Does your organisation conduct lifecycle assessments for its products and apply product stewardship principles across the full value chain?",
        'tr': "Kuruluşunuz ürünleri için yaşam döngüsü değerlendirmeleri yapıyor ve tüm değer zinciri boyunca ürün gözetim ilkelerini uyguluyor mu?",
    },
    'MFG-7': {
        'en': "Does your organisation measure and actively reduce GHG emissions from manufacturing processes, with science-based reduction targets?",
        'tr': "Kuruluşunuz bilim temelli azaltma hedefleriyle imalat süreçlerinden kaynaklanan sera gazı emisyonlarını ölçüyor ve aktif olarak azaltıyor mu?",
    },
    'MFG-8': {
        'en': "Does your organisation assess and manage the social and environmental impact of its operations on local communities?",
        'tr': "Kuruluşunuz faaliyetlerinin yerel topluluklar üzerindeki sosyal ve çevresel etkisini değerlendiriyor ve yönetiyor mu?",
    },

    # ── Construction & Real Estate (v4) ────────────────────────────────────
    'CON-1': {
        'en': "Does your organisation design and certify new buildings to recognised green building standards (e.g. LEED Gold+, BREEAM Excellent+)?",
        'tr': "Kuruluşunuz yeni binaları tanınan yeşil bina standartlarına (LEED Gold+, BREEAM Excellent+ vb.) göre tasarlayıp sertifikalandırıyor mu?",
    },
    'CON-2': {
        'en': "Does your organisation measure and minimise embodied carbon in construction materials across all new projects?",
        'tr': "Kuruluşunuz tüm yeni projelerde inşaat malzemelerindeki gömülü karbonu ölçüyor ve en aza indiriyor mu?",
    },
    'CON-3': {
        'en': "Does your organisation assess and mitigate the biodiversity impact of development projects, and aim for biodiversity net gain?",
        'tr': "Kuruluşunuz geliştirme projelerinin biyoçeşitlilik üzerindeki etkisini değerlendiriyor, azaltıyor ve biyoçeşitlilik net kazanımı hedefliyor mu?",
    },
    'CON-4': {
        'en': "Does your organisation have targets and systems to reduce, reuse and recycle construction and demolition waste?",
        'tr': "Kuruluşunuz inşaat ve yıkım atıklarını azaltmaya, yeniden kullanmaya ve geri dönüştürmeye yönelik hedef ve sistemlere sahip mi?",
    },
    'CON-5': {
        'en': "Does your organisation have a comprehensive health and safety management system for all workers on construction sites?",
        'tr': "Kuruluşunuz şantiyedeki tüm çalışanlar için kapsamlı bir iş sağlığı ve güvenliği yönetim sistemine sahip mi?",
    },
    'CON-6': {
        'en': "Does your organisation contribute to affordable housing and assess community needs as part of its development planning process?",
        'tr': "Kuruluşunuz geliştirme planlama sürecinde uygun fiyatlı konut üretimine katkıda bulunuyor ve toplumsal ihtiyaçları değerlendiriyor mu?",
    },
    'CON-7': {
        'en': "Does your organisation incorporate climate resilience into the design and long-term management of built assets?",
        'tr': "Kuruluşunuz inşa edilen varlıkların tasarımına ve uzun vadeli yönetimine iklim dayanıklılığını dahil ediyor mu?",
    },
    'CON-8': {
        'en': "Does your organisation have a policy to respect and protect the land and resource rights of communities affected by its developments?",
        'tr': "Kuruluşunuzun geliştirmelerinden etkilenen toplulukların arazi ve kaynak haklarını saygıyla korumaya yönelik bir politikası var mı?",
    },

    # ── Healthcare & Pharma (v4) ───────────────────────────────────────────
    'HLT-1': {
        'en': "Does your organisation have a published access-to-medicines strategy with tiered pricing, voluntary licensing and outcome targets?",
        'tr': "Kuruluşunuzun kademeli fiyatlandırma, gönüllü lisanslama ve sonuç hedeflerini içeren yayımlanmış bir ilaca erişim stratejisi var mı?",
    },
    'HLT-2': {
        'en': "Does your organisation register and publicly disclose results of all clinical trials, including negative findings, within 12 months of completion?",
        'tr': "Kuruluşunuz tüm klinik deneyleri kaydettiriyor ve tamamlanmadan 12 ay içinde olumsuz bulgular dahil sonuçları kamuoyuyla paylaşıyor mu?",
    },
    'HLT-3': {
        'en': "Does your organisation have a programme to combat antimicrobial resistance through responsible use, stewardship and R&D investment?",
        'tr': "Kuruluşunuzun sorumlu kullanım, gözetim ve Ar-Ge yatırımı yoluyla antimikrobiyal dirençle mücadeleye yönelik bir programı var mı?",
    },
    'HLT-4': {
        'en': "Does your organisation have a management system for the safe disposal of medical and hazardous waste that exceeds regulatory minimums?",
        'tr': "Kuruluşunuzun yasal asgari gerekliliklerin üzerinde tıbbi ve tehlikeli atıkların güvenli bertarafı için bir yönetim sistemi var mı?",
    },
    'HLT-5': {
        'en': "Does your organisation have a patient privacy and health data protection framework that is compliant with applicable regulations?",
        'tr': "Kuruluşunuzun yürürlükteki mevzuata uygun bir hasta gizliliği ve sağlık verisi koruma çerçevesi var mı?",
    },
    'HLT-6': {
        'en': "Does your organisation monitor and protect the health and safety of healthcare workers, including exposure to hazardous substances?",
        'tr': "Kuruluşunuz tehlikeli maddelere maruziyeti de kapsayacak şekilde sağlık çalışanlarının sağlığını ve güvenliğini izliyor ve koruyor mu?",
    },
    'HLT-7': {
        'en': "Does your organisation have a robust quality management system to ensure the safety and efficacy of all its products?",
        'tr': "Kuruluşunuzun tüm ürünlerinin güvenliğini ve etkinliğini sağlamaya yönelik sağlam bir kalite yönetim sistemi var mı?",
    },
    'HLT-8': {
        'en': "Does your organisation have and enforce an ethical marketing and promotion policy with independent audit and published breach data?",
        'tr': "Kuruluşunuzun bağımsız denetim ve yayımlanmış ihlal verileriyle desteklenen, uygulanan bir etik pazarlama ve tanıtım politikası var mı?",
    },

    # ── Technology & IT (v4) ───────────────────────────────────────────────
    'TECH-1': {
        'en': "Does your organisation have a comprehensive data privacy and security management framework with regular independent audits?",
        'tr': "Kuruluşunuzun düzenli bağımsız denetimlerle desteklenen kapsamlı bir veri gizliliği ve güvenlik yönetimi çerçevesi var mı?",
    },
    'TECH-2': {
        'en': "Does your organisation have governance processes for algorithmic accountability, including bias audits and transparency reporting?",
        'tr': "Kuruluşunuzun önyargı denetimi ve şeffaflık raporlamasını kapsayan algoritmik hesap verebilirliğe yönelik yönetim süreçleri var mı?",
    },
    'TECH-3': {
        'en': "Does your organisation operate a product take-back programme to responsibly manage e-waste across its product portfolio?",
        'tr': "Kuruluşunuz ürün portföyü genelinde e-atıkları sorumlu bir şekilde yönetmek için ürün geri alma programı yürütüyor mu?",
    },
    'TECH-4': {
        'en': "Does your organisation measure and actively reduce the energy consumption of its data centres and digital infrastructure?",
        'tr': "Kuruluşunuz veri merkezlerinin ve dijital altyapısının enerji tüketimini ölçüyor ve aktif olarak azaltıyor mu?",
    },
    'TECH-5': {
        'en': "Does your organisation have programmes to promote digital inclusion and ensure equitable access to its products and services?",
        'tr': "Kuruluşunuz dijital kapsayıcılığı teşvik etmeye ve ürün ile hizmetlerine eşit erişimi sağlamaya yönelik programlara sahip mi?",
    },
    'TECH-6': {
        'en': "Does your organisation have a published responsible AI policy that addresses ethics, bias, human oversight and the impact of automation on workers?",
        'tr': "Kuruluşunuzun etik, önyargı, insan gözetimi ve otomasyonun çalışanlar üzerindeki etkisini ele alan yayımlanmış bir sorumlu yapay zeka politikası var mı?",
    },
    'TECH-7': {
        'en': "Does your organisation conduct OECD-aligned due diligence on conflict minerals in its supply chain with independent smelter audits?",
        'tr': "Kuruluşunuz tedarik zincirinde çatışma mineralleri üzerinde bağımsız eritme tesisi denetimleriyle desteklenen OECD uyumlu özeni yürütüyor mu?",
    },
    'TECH-8': {
        'en': "Does your organisation have a comprehensive cybersecurity risk management and incident response framework with published performance metrics?",
        'tr': "Kuruluşunuzun yayımlanmış performans metrikleriyle desteklenen kapsamlı bir siber güvenlik risk yönetimi ve olay müdahale çerçevesi var mı?",
    },

    # ── Retail & Trade (v4) ────────────────────────────────────────────────
    'RET-1': {
        'en': "Does your organisation have a policy and monitoring system to ensure fair labour standards throughout its supply chain?",
        'tr': "Kuruluşunuzun tedarik zinciri boyunca adil çalışma standartlarını sağlamaya yönelik bir politikası ve izleme sistemi var mı?",
    },
    'RET-2': {
        'en': "Does your organisation have a product safety management system and a documented product recall procedure with published safety data?",
        'tr': "Kuruluşunuzun yayımlanmış güvenlik verileriyle desteklenen bir ürün güvenliği yönetim sistemi ve belgelenmiş ürün geri çağırma prosedürü var mı?",
    },
    'RET-3': {
        'en': "Does your organisation have targets to reduce packaging, increase recyclability and eliminate unnecessary single-use plastic?",
        'tr': "Kuruluşunuzun ambalajı azaltmaya, geri dönüştürülebilirliği artırmaya ve gereksiz tek kullanımlık plastikleri ortadan kaldırmaya yönelik hedefleri var mı?",
    },
    'RET-4': {
        'en': "Does your organisation have a consumer data privacy and protection framework compliant with applicable data protection regulations?",
        'tr': "Kuruluşunuzun geçerli veri koruma yönetmeliklerine uygun bir tüketici veri gizliliği ve koruma çerçevesi var mı?",
    },
    'RET-5': {
        'en': "Does your organisation have a responsible sourcing policy for raw materials that addresses environmental and social risks in its supply chain?",
        'tr': "Kuruluşunuzun tedarik zincirindeki çevresel ve sosyal riskleri ele alan ham maddeler için sorumlu tedarik politikası var mı?",
    },
    'RET-6': {
        'en': "Does your organisation measure and actively reduce food waste across its retail operations, with published reduction targets?",
        'tr': "Kuruluşunuz yayımlanmış azaltma hedefleriyle perakende operasyonları genelinde gıda israfını ölçüyor ve aktif olarak azaltıyor mu?",
    },
    'RET-7': {
        'en': "Does your organisation have a commitment to paying living wages to all workers in its supply chain, with independent verification?",
        'tr': "Kuruluşunuzun bağımsız doğrulamayla birlikte tedarik zincirindeki tüm çalışanlara yaşanabilir ücret ödemeye yönelik bir taahhüdü var mı?",
    },
    'RET-8': {
        'en': "Does your organisation measure and reduce the carbon footprint of its logistics and last-mile delivery operations?",
        'tr': "Kuruluşunuz lojistik ve son kilometre teslimat operasyonlarının karbon ayak izini ölçüyor ve azaltıyor mu?",
    },

    # ════════════════════════════════════════════════════════════════════════
    # v3 legacy IDs  (AG-01 … RET-08)  — kept for backward compatibility
    # ════════════════════════════════════════════════════════════════════════

    # ── Agriculture & Food (v3) ────────────────────────────────────────────
    'AG-01': {
        'en': "What percentage of your sourcing volume is covered by third-party sustainability certifications (e.g. RSPO, FSC, Rainforest Alliance)?",
        'tr': "Kaynak hacminizin yüzde kaçı üçüncü taraf sürdürülebilirlik sertifikalarıyla (RSPO, FSC, Rainforest Alliance vb.) kapsanmaktadır?",
    },
    'AG-02': {
        'en': "Does your organisation have a verified zero-deforestation commitment with satellite monitoring?",
        'tr': "Kuruluşunuzun uydu izleme ile doğrulanmış sıfır ormansızlaştırma taahhüdü var mı?",
    },
    'AG-03': {
        'en': "How does your organisation manage pesticide use and implement Integrated Pest Management (IPM)?",
        'tr': "Kuruluşunuz pestisit kullanımını ve Entegre Zararlı Yönetimini (IPM) nasıl yönetmektedir?",
    },
    'AG-04': {
        'en': "Does your organisation track soil health and water efficiency, and adopt regenerative agriculture practices?",
        'tr': "Kuruluşunuz toprak sağlığını ve su verimliliğini izliyor, yenileyici tarım uygulamalarını benimsiyor mu?",
    },
    'AG-05': {
        'en': "What income, training and financing support programmes does your organisation provide to smallholder farmers?",
        'tr': "Kuruluşunuz küçük çiftçilere yönelik gelir, eğitim ve finansman destek programları sunuyor mu?",
    },
    'AG-06': {
        'en': "How does your organisation ensure and report on animal welfare standards?",
        'tr': "Kuruluşunuz hayvan refahı standartlarını nasıl sağlamakta ve raporlamaktadır?",
    },
    'AG-07': {
        'en': "Does your organisation measure food waste across all stages and have a reduction target aligned with SDG 12.3?",
        'tr': "Kuruluşunuz gıda israfını tüm aşamalarda ölçüyor mu ve SDG 12.3 ile uyumlu azaltma hedefi var mı?",
    },
    'AG-08': {
        'en': "Does your organisation collect farmer income data and apply a living income benchmark?",
        'tr': "Kuruluşunuz çiftçi gelir verisi toplayarak geçim standardı gelir karşılaştırması uyguluyor mu?",
    },

    # ── Energy & Utilities ─────────────────────────────────────────────────
    'EN-01': {
        'en': "Does your organisation have a published Just Transition plan with retraining budget and community investment commitments?",
        'tr': "Kuruluşunuzun yeniden eğitim bütçesi ve toplumsal yatırım taahhütlerini içeren yayımlanmış bir Adil Geçiş planı var mı?",
    },
    'EN-02': {
        'en': "Are your organisation's asset retirement and decommissioning obligations fully funded and publicly disclosed?",
        'tr': "Kuruluşunuzun varlık emeklilik/devreden çıkarma yükümlülükleri tam olarak karşılanmış ve kamuoyuyla paylaşılmış mı?",
    },
    'EN-03': {
        'en': "How does your organisation measure, manage and reduce methane emissions?",
        'tr': "Kuruluşunuz metan emisyonlarını nasıl ölçmekte, yönetmekte ve azaltmaktadır?",
    },
    'EN-04': {
        'en': "What percentage of your organisation's generation or sales capacity is zero-carbon?",
        'tr': "Kuruluşunuzun üretim veya satış kapasitesinin yüzde kaçı sıfır karbon kaynaklıdır?",
    },
    'EN-05': {
        'en': "Does your organisation have a quantified energy access programme for underserved communities with publicly disclosed outcomes?",
        'tr': "Kuruluşunuz hizmet götürülemeyen topluluklar için ölçülebilir ve kamuoyuyla paylaşılmış sonuçlara sahip bir enerji erişim programı yürütüyor mu?",
    },
    'EN-06': {
        'en': "Does your organisation track water intensity and have an improvement target versus a base year?",
        'tr': "Kuruluşunuz su yoğunluğunu izliyor mu ve baz yılına göre iyileştirme hedefi var mı?",
    },
    'EN-07': {
        'en': "How does your organisation measure and report air emissions (SO₂, NOₓ, PM) relative to Best Available Technology (BAT) standards?",
        'tr': "Kuruluşunuz hava emisyonlarını (SO₂, NOₓ, PM) En İyi Mevcut Teknoloji (BAT) standartlarına göre nasıl ölçmekte ve raporlamaktadır?",
    },
    'EN-08': {
        'en': "Does your organisation have a published climate-resilient infrastructure investment plan?",
        'tr': "Kuruluşunuzun yayımlanmış iklime dayanıklı altyapı yatırım planı var mı?",
    },

    # ── Financial Services ─────────────────────────────────────────────────
    'FIN-01': {
        'en': "Has your organisation conducted a full PCAF financed emissions inventory with third-party verification and set a reduction target?",
        'tr': "Kuruluşunuz üçüncü taraf doğrulamasıyla tam PCAF finanse edilen emisyon envanteri hazırladı ve azaltma hedefi belirledi mi?",
    },
    'FIN-02': {
        'en': "Does your organisation conduct TCFD scenario analysis (1.5°C and 3°C+) across credit and market portfolios?",
        'tr': "Kuruluşunuz kredi ve piyasa portföylerinde 1,5°C ve 3°C+ senaryoları için TCFD iklim riski analizi yapıyor mu?",
    },
    'FIN-03': {
        'en': "How does your organisation integrate ESG factors into investment and lending decisions?",
        'tr': "Kuruluşunuz ESG faktörlerini yatırım ve kredi kararlarına nasıl entegre etmektedir?",
    },
    'FIN-04': {
        'en': "Does your organisation have a published exclusion policy covering coal, controversial weapons and deforestation-linked sectors?",
        'tr': "Kuruluşunuzun kömür, tartışmalı silahlar ve ormansızlaştırmayla bağlantılı sektörleri kapsayan yayımlanmış bir dışlama politikası var mı?",
    },
    'FIN-05': {
        'en': "What percentage of your organisation's new lending or assets under management is aligned with sustainable finance taxonomies?",
        'tr': "Kuruluşunuzun yeni kredi hacmi veya yönetim altındaki varlıkların yüzde kaçı sürdürülebilir finans taksonomileriyle uyumludur?",
    },
    'FIN-06': {
        'en': "Does your organisation have a quantified financial inclusion programme with published outcomes?",
        'tr': "Kuruluşunuz yayımlanmış sonuçlarla birlikte ölçülebilir bir finansal kapsayıcılık programına sahip mi?",
    },
    'FIN-07': {
        'en': "How many notifiable data privacy breaches did your organisation experience in the reporting period?",
        'tr': "Raporlama döneminde kuruluşunuz kaç tane bildirilebilir veri gizliliği ihlali yaşadı?",
    },
    'FIN-08': {
        'en': "Does your organisation disclose nature-related risks (TNFD) and incorporate them into credit and investment decisions?",
        'tr': "Kuruluşunuz TNFD çerçevesiyle doğayla ilgili riskleri açıklıyor ve bunları kredi ve yatırım kararlarına dahil ediyor mu?",
    },

    # ── Manufacturing & Industry ───────────────────────────────────────────
    'MFG-01': {
        'en': "What percentage of revenue-weighted product lines has undergone a Life Cycle Assessment (LCA), and is eco-design mandatory for new products?",
        'tr': "Gelir ağırlıklı ürün hatlarınızın yüzde kaçı Yaşam Döngüsü Değerlendirmesi (LCA) geçirmiştir ve yeni ürünlerde eko-tasarım zorunlu mu?",
    },
    'MFG-02': {
        'en': "What percentage of your total material inputs (by weight) consists of recycled or reused content?",
        'tr': "Toplam malzeme girdilerinizin (ağırlık olarak) yüzde kaçı geri dönüştürülmüş veya yeniden kullanılmış içerikten oluşmaktadır?",
    },
    'MFG-03': {
        'en': "Does your organisation maintain a Substances of Very High Concern (SVHC) inventory and have an active substitution plan?",
        'tr': "Kuruluşunuz Çok Endişe Verici Madde (SVHC) envanteri tutuyor mu ve aktif bir ikame planı var mı?",
    },
    'MFG-04': {
        'en': "How does your organisation's energy intensity compare to industry best-in-class benchmarks?",
        'tr': "Kuruluşunuzun enerji yoğunluğu sektördeki en iyi uygulamalarla karşılaştırıldığında nasıl bir konumdadır?",
    },
    'MFG-05': {
        'en': "Does your organisation participate in an Extended Producer Responsibility (EPR) scheme and publicly report collection rates?",
        'tr': "Kuruluşunuz Genişletilmiş Üretici Sorumluluğu (EPR) programına katılıyor ve toplama oranlarını kamuoyuyla paylaşıyor mu?",
    },
    'MFG-06': {
        'en': "Has your organisation mapped its Tier 2 suppliers and established active improvement plans?",
        'tr': "Kuruluşunuz 2. Kademe tedarikçilerini haritalandırdı ve aktif iyileştirme planları oluşturdu mu?",
    },
    'MFG-07': {
        'en': "How does your organisation track, report and share learnings from Tier 1 Process Safety Events (PSEs)?",
        'tr': "Kuruluşunuz 1. Kademe Proses Güvenliği Olaylarını (PSE) nasıl takip etmekte, raporlamakta ve sektörle paylaşmaktadır?",
    },
    'MFG-08': {
        'en': "Does your organisation track water intensity in manufacturing and have an improvement target versus a base year?",
        'tr': "Kuruluşunuz imalattaki su yoğunluğunu izliyor mu ve baz yılına göre iyileştirme hedefi var mı?",
    },

    # ── Construction & Real Estate ─────────────────────────────────────────
    'CON-01': {
        'en': "What percentage of your property portfolio holds a LEED Gold+, BREEAM Excellent+ or equivalent green building certification?",
        'tr': "Mülk portföyünüzün yüzde kaçı LEED Gold+, BREEAM Excellent+ veya eşdeğeri yeşil bina sertifikasına sahiptir?",
    },
    'CON-02': {
        'en': "Does your organisation assess whole-life embodied carbon for all new projects and set low-carbon materials targets?",
        'tr': "Kuruluşunuz tüm yeni projelerde tüm yaşam döngüsü boyunca gömülü karbonu değerlendiriyor ve düşük karbonlu malzeme hedefleri belirliyor mu?",
    },
    'CON-03': {
        'en': "What percentage of your portfolio has an Energy Performance Certificate (EPC) rating of B or above?",
        'tr': "Portföyünüzün yüzde kaçı B veya üzeri Enerji Performans Sertifikası (EPC) derecesine sahiptir?",
    },
    'CON-04': {
        'en': "Has your organisation assessed physical climate risks across its entire portfolio and integrated findings into asset valuations?",
        'tr': "Kuruluşunuz portföyünün tamamında fiziksel iklim risklerini değerlendirdi ve bulguları varlık değerlemelerine entegre etti mi?",
    },
    'CON-05': {
        'en': "What percentage of construction and demolition waste is diverted from landfill across your projects?",
        'tr': "Projelerinizde inşaat ve yıkım atıklarının yüzde kaçı düzenli depolama alanlarından uzak tutulmaktadır?",
    },
    'CON-06': {
        'en': "Does your organisation achieve and verify Biodiversity Net Gain (BNG) of at least 10% on all new developments?",
        'tr': "Kuruluşunuz tüm yeni gelişmelerde en az %10 Biyoçeşitlilik Net Kazanımı (BNG) sağlıyor ve doğruluyor mu?",
    },
    'CON-07': {
        'en': "Does your organisation quantify social value (e.g. SROI) for all major developments?",
        'tr': "Kuruluşunuz tüm büyük gelişmeler için sosyal değeri (örn. SROI) ölçüyor mu?",
    },
    'CON-08': {
        'en': "Does your organisation have a structured green lease programme and collect tenant energy and waste data?",
        'tr': "Kuruluşunuzun yapılandırılmış bir yeşil kira programı var mı ve kiracı enerji ile atık verilerini topluyor mu?",
    },

    # ── Healthcare & Pharma ────────────────────────────────────────────────
    'HC-01': {
        'en': "Does your organisation have a published Access to Medicine strategy with tiered pricing and voluntary licensing commitments?",
        'tr': "Kuruluşunuzun kademeli fiyatlandırma ve gönüllü lisanslama taahhütlerini içeren yayımlanmış bir İlaca Erişim stratejisi var mı?",
    },
    'HC-02': {
        'en': "Is your organisation's pharmacovigilance system certified and free of critical regulatory findings for the past three years?",
        'tr': "Kuruluşunuzun farmakovijilans sistemi sertifikalı mı ve son üç yılda kritik düzenleyici bulgu yoktur mu?",
    },
    'HC-03': {
        'en': "Does your organisation register 100% of clinical trials and publish results within 12 months of completion?",
        'tr': "Kuruluşunuz klinik deneylerin %100'ünü kaydettiriyor ve sonuçları tamamlanmadan 12 ay içinde yayımlıyor mu?",
    },
    'HC-04': {
        'en': "Has your organisation signed the IFPMA Antimicrobial Resistance (AMR) commitment and implemented active antimicrobial stewardship?",
        'tr': "Kuruluşunuz IFPMA Antimikrobiyal Direnç (AMR) taahhüdünü imzaladı ve aktif antimikrobiyal koruyuculuk programı uyguluyor mu?",
    },
    'HC-05': {
        'en': "Does your organisation treat pharmaceutical waste above regulatory minimums and undergo third-party audits?",
        'tr': "Kuruluşunuz ilaç atıklarını yasal asgari düzeyin üzerinde arıtıyor ve üçüncü taraf denetimine tabi tutuyor mu?",
    },
    'HC-06': {
        'en': "Does your organisation adhere to the IFPMA marketing ethics code with independent audit and published breach data?",
        'tr': "Kuruluşunuz bağımsız denetim ve yayımlanan ihlal verileriyle birlikte IFPMA pazarlama etiği koduna uyuyor mu?",
    },
    'HC-07': {
        'en': "Does your organisation provide Continuing Professional Development (CPD) above regulatory requirements and publicly track hours?",
        'tr': "Kuruluşunuz yasal gerekliliklerin üzerinde Sürekli Mesleki Gelişim (SMG) sunuyor ve saatleri kamuoyuyla paylaşıyor mu?",
    },
    'HC-08': {
        'en': "Is your organisation's pharmaceutical supply chain GDP-compliant across all tiers with an active audit programme?",
        'tr': "Kuruluşunuzun ilaç tedarik zinciri tüm kademelerde İyi Dağıtım Uygulamaları (GDP) ile uyumlu ve aktif denetim programına sahip mi?",
    },

    # ── Technology & IT ────────────────────────────────────────────────────
    'TECH-01': {
        'en': "What is your organisation's average data centre Power Usage Effectiveness (PUE) and what percentage is powered by renewable energy?",
        'tr': "Kuruluşunuzun veri merkezlerinin ortalama Güç Kullanım Etkinliği (PUE) değeri nedir ve yüzde kaçı yenilenebilir enerjiyle çalışmaktadır?",
    },
    'TECH-02': {
        'en': "Has your organisation quantified Scope 3 emissions from product use-phase and end-of-life stages?",
        'tr': "Kuruluşunuz ürünlerin kullanım aşaması ve kullanım ömrü sonu Kapsam 3 emisyonlarını ölçüp ölçmediğini belirledi mi?",
    },
    'TECH-03': {
        'en': "Does your organisation have an e-waste take-back programme and what percentage of devices are refurbished or recycled?",
        'tr': "Kuruluşunuzun e-atık geri alma programı var mı ve cihazların yüzde kaçı yeniden işleniyor veya geri dönüştürülüyor?",
    },
    'TECH-04': {
        'en': "Is your organisation ISO 27001 certified with annual penetration testing and zero notifiable breaches in the past three years?",
        'tr': "Kuruluşunuz ISO 27001 sertifikasına sahip mi, yıllık sızma testi yapıyor mu ve son üç yılda bildirilebilir ihlal sayısı sıfır mı?",
    },
    'TECH-05': {
        'en': "Does your organisation have a published AI ethics policy with algorithmic bias audits and human oversight for high-risk applications?",
        'tr': "Kuruluşunuzun algoritmik önyargı denetimi ve yüksek riskli uygulamalar için insan gözetimini kapsayan yayımlanmış bir Yapay Zeka Etiği politikası var mı?",
    },
    'TECH-06': {
        'en': "Do your organisation's products meet WCAG 2.1 AA accessibility standards and include an affordability programme for underserved populations?",
        'tr': "Kuruluşunuzun ürünleri WCAG 2.1 AA erişilebilirlik standartlarını karşılıyor mu ve hizmet götürülemeyen kesimlere yönelik erişilebilirlik programı sunuyor mu?",
    },
    'TECH-07': {
        'en': "Does your organisation conduct OECD due diligence for all 3TG and cobalt minerals with a smelter audit programme?",
        'tr': "Kuruluşunuz eritme tesisi denetim programıyla birlikte tüm 3TG ve kobalt mineralleri için OECD kaynaklı mineral özeni yürütüyor mu?",
    },
    'TECH-08': {
        'en': "Does your organisation quantify employee home-working carbon emissions and provide a low-carbon home-working support programme?",
        'tr': "Kuruluşunuz çalışanların evden çalışma kaynaklı karbon emisyonlarını ölçüyor ve düşük karbonlu evden çalışma destek programı sunuyor mu?",
    },

    # ── Retail & Trade ─────────────────────────────────────────────────────
    'RET-01': {
        'en': "Does your organisation publish sustainable product standards, and what percentage of revenue comes from products meeting those standards?",
        'tr': "Kuruluşunuz sürdürülebilir ürün standartları yayımlıyor mu ve bu standartları karşılayan ürünlerden elde edilen gelir yüzdesi ne kadardır?",
    },
    'RET-02': {
        'en': "What percentage of all packaging is recyclable, reusable or compostable, and does your organisation have a plastic reduction target?",
        'tr': "Tüm ambalajların yüzde kaçı geri dönüştürülebilir, yeniden kullanılabilir veya kompostlanabilir; plastik azaltma hedefiniz var mı?",
    },
    'RET-03': {
        'en': "How does your organisation audit and assess labour standards across its supply chain tiers?",
        'tr': "Kuruluşunuz tedarik zinciri kademeleri genelindeki çalışma koşullarını nasıl denetlemekte ve değerlendirmektedir?",
    },
    'RET-04': {
        'en': "Does your organisation measure food waste across all stages, have a reduction target and operate a surplus food donation programme?",
        'tr': "Kuruluşunuz gıda israfını tüm aşamalarda ölçüyor, azaltma hedefi belirliyor ve fazla gıda bağış programı yürütüyor mu?",
    },
    'RET-05': {
        'en': "Does your organisation have product reformulation targets for salt, sugar and fat, and apply front-of-pack nutrition labelling?",
        'tr': "Kuruluşunuzun tuz, şeker ve yağ için ürün reformülasyon hedefleri var mı ve ön etiket beslenme bilgisi sunuyor mu?",
    },
    'RET-06': {
        'en': "Does your organisation quantify last-mile logistics emissions and have a fleet electrification transition plan?",
        'tr': "Kuruluşunuz son kilometre lojistik emisyonlarını ölçüyor mu ve filo elektrifikasyon geçiş planına sahip mi?",
    },
    'RET-07': {
        'en': "Does your organisation operate a take-back, resale or repair programme and track the percentage of units diverted from landfill?",
        'tr': "Kuruluşunuz ürün geri alma, yeniden satış veya tamir programı işletiyor ve düzenli depolama alanından uzak tutulan ürün yüzdesini takip ediyor mu?",
    },
    'RET-08': {
        'en': "Has your organisation extended a living wage commitment to Tier 1 supply chain workers with independent verification?",
        'tr': "Kuruluşunuz bağımsız doğrulamayla birlikte 1. Kademe tedarik zinciri çalışanlarına yaşanabilir ücret taahhüdünü genişletti mi?",
    },
}


# ── Helpers ────────────────────────────────────────────────────────────────

def _s(v):
    if v is None: return ''
    return str(v).strip()

def _n(v, d=0):
    # Fix L-5: bare except: catches SystemExit/KeyboardInterrupt — use specific types.
    try: return int(float(v))
    except (TypeError, ValueError): return d

def _find_header(ws, col_b_val):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if _s(row[1]) == col_b_val:
            return i
    return None


# ══════════════════════════════════════════════════════════════════════════
class Command(BaseCommand):
    help = 'Import GRI v3 Excel — creates one Survey per section (12 total).'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str)
        parser.add_argument('--clear', action='store_true',
                            help='Delete all existing GRI surveys first')

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl not installed. Run: pip install openpyxl==3.1.2')

        path = options['excel_path']
        self.stdout.write(f'[GRI] Loading: {path}')
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {path}')
        except Exception as e:
            raise CommandError(f'Cannot open file: {e}')

        # Fix CRASH-01: always close the read-only workbook before the process
        # exits.  openpyxl's read_only mode keeps a ZipFile (and its underlying
        # I/O threads) open for the entire workbook lifetime.  Without an
        # explicit close(), Python's garbage collector tries to finalize the
        # ZipFile during interpreter shutdown, which races against the Kerberos
        # pthread mutex teardown on Render's Linux image and triggers the
        # "k5_mutex_lock: Assertion `r == 0' failed" abort (core dump).
        # Wrapping the entire body in try/finally guarantees wb.close() is
        # called even when an exception is raised mid-import.
        try:
            # -- Optional clear --
            if options['clear']:
                n, _ = Survey.objects.filter(name__istartswith='GRI').delete()
                self.stdout.write(self.style.WARNING(f'  [!] Deleted {n} existing GRI surveys.'))

            total_surveys = total_q = total_c = 0

            # -- Core sections --
            self.stdout.write('\n-- Core sections -------------------------------------------')
            for cfg in CORE_SECTIONS:
                if cfg['sheet'] not in wb.sheetnames:
                    self.stdout.write(self.style.ERROR(f'  [X] Sheet missing: {cfg["sheet"]}'))
                    continue

                survey = self._upsert_survey(
                    cfg['survey_name'], cfg['survey_name_tr'], cfg['survey_desc']
                )
                cat = self._upsert_category(
                    survey, cfg['cat_name'], cfg['cat_name_tr'],
                    order=1,
                    max_score=cfg['max_score'],
                    env_w=cfg['env_w'], soc_w=cfg['soc_w'], gov_w=cfg['gov_w'],
                )
                questions = self._parse_core(wb[cfg['sheet']])
                q, c = self._persist(questions, survey, cat)
                self._fix_score_anomalies(survey)
                total_surveys += 1; total_q += q; total_c += c
                self.stdout.write(f'  [OK] {cfg["survey_name"]:<45} {q:>3}Q  {c:>4}C')

            # -- Sector modules --
            self.stdout.write('\n-- Sector modules ------------------------------------------')
            for cfg in SECTOR_SECTIONS:
                if cfg['sheet'] not in wb.sheetnames:
                    self.stdout.write(self.style.WARNING(f'  [!] Sheet missing: {cfg["sheet"]}'))
                    continue

                survey = self._upsert_survey(
                    cfg['survey_name'], cfg['survey_name_tr'], cfg['survey_desc']
                )
                cat = self._upsert_category(
                    survey, cfg['cat_name'], cfg['cat_name_tr'],
                    order=1, max_score=80,
                    env_w=0.34, soc_w=0.33, gov_w=0.33,
                )
                questions = self._parse_sector(wb[cfg['sheet']])
                q, c = self._persist(questions, survey, cat)
                total_surveys += 1; total_q += q; total_c += c
                self.stdout.write(f'  [OK] {cfg["survey_name"]:<45} {q:>3}Q  {c:>4}C')

            self.stdout.write(self.style.SUCCESS(
                f'\n[DONE] {total_surveys} surveys, {total_q} questions, {total_c} choices imported.\n'
            ))
        finally:
            wb.close()

    # ── Survey / Category upsert ───────────────────────────────────────────

    def _upsert_survey(self, name, name_tr, desc):
        survey, _ = Survey.objects.update_or_create(
            name=name,
            defaults={
                'name_en': name,
                'name_tr': name_tr,
                'description': desc,
                'description_en': desc,
                'description_tr': desc,
                'is_active': True,
                'allow_multiple_attempts': True,
                'show_results_immediately': True,
            }
        )
        return survey

    def _upsert_category(self, survey, name, name_tr, order,
                         max_score, env_w, soc_w, gov_w):
        cat, _ = Category.objects.update_or_create(
            survey=survey, name=name,
            defaults={
                'name_en': name, 'name_tr': name_tr,
                'order': order, 'max_score': max_score,
                'environmental_weight': env_w,
                'social_weight': soc_w,
                'governance_weight': gov_w,
            }
        )
        return cat

    # ── Parsers ────────────────────────────────────────────────────────────

    def _parse_core(self, ws):
        hr = _find_header(ws, 'ID')
        if not hr: return {}
        questions = {}
        cur_id = None
        order = 0
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            p = list(row) + [None] * 12
            q_id, layer_name, q_text, option, pts = _s(p[1]), _s(p[3]), _s(p[5]), _s(p[7]), _n(p[8])
            if not option or option[0] not in ('A', 'B', 'C', 'D'):
                continue
            if q_id and q_id not in ('ID', '#'):
                if q_id not in questions:
                    order += 1
                    # Store clean question text: no [q_id] code prefix.
                    # The layer name is appended in brackets so GRI assessors
                    # can identify which Policy/Implementation/Measurement/Results
                    # layer the question belongs to.  The frontend strips both
                    # the code prefix (already removed here) and the layer suffix
                    # via cleanQuestionText() when rendering to end users.
                    label = q_text if q_text else q_id
                    if layer_name:
                        label = f'{label}  [{layer_name}]'
                    questions[q_id] = {'text': label, 'order': order, 'choices': []}
                cur_id = q_id
            if cur_id and cur_id in questions:
                questions[cur_id]['choices'].append({
                    'text': option, 'score': pts,
                    'order': LETTER_ORDER.get(option[0], 99)
                })
        return questions

    def _parse_sector(self, ws):
        hr = _find_header(ws, 'Q ID')
        if not hr: return {}
        questions = {}
        cur_id = None
        order = 0
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            p = list(row) + [None] * 12
            q_id    = _s(p[1])
            cat_nm  = _s(p[2])          # Category column   (C)
            q_text  = _s(p[3])          # Question col  (D) — v3; v4 puts "Sector" here
            gri_ref = _s(p[4])          # GRI Ref col   (E) — v3; v4 uses Criterion here
            option  = _s(p[5])          # Answer Option (F)
            pts     = _n(p[6])          # Points        (G)
            if not option or option[0] not in ('A', 'B', 'C', 'D'):
                continue
            if q_id and q_id not in ('Q ID', '#'):
                if q_id not in questions:
                    order += 1
                    # v4 STRUCTURED puts a layer marker (e.g. "Sector") in the
                    # column that v3 used for the full question text.  Treat any
                    # LAYER_MARKERS value as absent so SECTOR_QUESTION_TEXT is
                    # consulted instead of storing the word "Sector" verbatim.
                    effective_text = q_text if q_text and q_text.lower() not in LAYER_MARKERS else ''

                    # Priority: real Excel text → SECTOR_QUESTION_TEXT → fallback
                    override = SECTOR_QUESTION_TEXT.get(q_id, {})
                    if effective_text:
                        label    = effective_text
                        label_tr = ''
                    elif override:
                        label    = override.get('en', cat_nm or q_id)
                        label_tr = override.get('tr', '')
                    elif cat_nm and gri_ref:
                        label    = f'{cat_nm}  ({gri_ref})'
                        label_tr = ''
                    else:
                        label    = cat_nm if cat_nm else q_id
                        label_tr = ''
                    questions[q_id] = {
                        'text': label, 'text_tr': label_tr,
                        'order': order, 'choices': [],
                    }
                cur_id = q_id
            if cur_id and cur_id in questions:
                questions[cur_id]['choices'].append({
                    'text': option, 'score': pts,
                    'order': LETTER_ORDER.get(option[0], 99)
                })
        return questions

    # ── Persist ────────────────────────────────────────────────────────────

    def _persist(self, questions, survey, category):
        q_count = c_count = 0
        for data in questions.values():
            if not data['choices']:
                continue
            # text_tr is provided by _parse_sector (via SECTOR_QUESTION_TEXT).
            # _parse_core leaves it absent from the dict; default to ''.
            text_tr = data.get('text_tr', '')
            q, new_q = Question.objects.update_or_create(
                survey=survey, category=category, order=data['order'],
                defaults={
                    'text': data['text'], 'text_en': data['text'], 'text_tr': text_tr,
                    'question_type': 'choice', 'is_active': True, 'allow_multiple': False,
                }
            )
            if new_q: q_count += 1
            for ch in sorted(data['choices'], key=lambda x: x['order']):
                _, new_c = Choice.objects.update_or_create(
                    question=q, order=ch['order'],
                    defaults={'text': ch['text'], 'text_en': ch['text'], 'text_tr': '', 'score': ch['score']}
                )
                if new_c: c_count += 1
        return q_count, c_count

    # ── Fix score ordering anomalies ───────────────────────────────────────

    def _fix_score_anomalies(self, survey):
        """Ensure choice scores descend strictly A ≥ B ≥ C ≥ D.

        Re-sorts the scores in descending order across all choices (ordered
        by the LETTER_ORDER key: A=1, B=2, C=3, D=4) so that the best
        answer (A) always carries the highest score and the worst (D) the
        lowest.  The old implementation only swapped A and B, missing
        anomalies like C > B or D > C.
        """
        fixed = 0
        for q in Question.objects.filter(survey=survey).prefetch_related('choices'):
            # Sort choices by their display order (A=1, B=2, …)
            ordered = sorted(q.choices.all(), key=lambda c: c.order)
            if len(ordered) < 2:
                continue
            original = [c.score for c in ordered]
            corrected = sorted(original, reverse=True)   # A gets max, D gets min
            if original != corrected:
                for choice, new_score in zip(ordered, corrected):
                    if choice.score != new_score:
                        Choice.objects.filter(pk=choice.pk).update(score=new_score)
                fixed += 1
        return fixed
