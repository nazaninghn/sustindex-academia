"""
Generate GRI_Questionnaire_v4_STRUCTURED.xlsx

Structure
---------
  Sheet 1: "GRI 1 — Foundation"          8 criteria × 4 layers = 32 questions
  Sheet 2: "GRI 2 — General Disclosures" 20 criteria × 4 layers = 80 questions
  Sheet 3: "GRI 3 — Material Topics"     15 criteria × 4 layers = 60 questions
  Sheets 4-11: 8 sector sheets (same format as v3)

Each question has 4 answer choices A/B/C/D scored 5/3/1/0.

Run: python generate_gri_v4.py
Output: data/GRI_Questionnaire_v4_STRUCTURED.xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT_PATH = os.path.join(os.path.dirname(__file__), "data", "GRI_Questionnaire_v4_STRUCTURED.xlsx")

# ── Scoring rubric ────────────────────────────────────────────────────────────
# A=5, B=3, C=1, D=0 per GRI maturity scale

def _choices_policy(topic):
    return [
        (f"A: Formally documented, board-approved policy on {topic} with measurable targets and scheduled review cycle.", 5),
        (f"B: Written policy on {topic} exists but lacks board approval, measurable targets, or regular review.", 3),
        (f"C: Informal or verbal commitment to {topic}; no formal documentation or targets.", 1),
        (f"D: No policy or commitment on {topic} in place.", 0),
    ]

def _choices_impl(topic):
    return [
        (f"A: {topic} fully implemented across all operations; responsible staff trained and accountable owners assigned.", 5),
        (f"B: {topic} partially implemented; gaps exist in coverage, staff training, or accountability.", 3),
        (f"C: {topic} addressed only in specific departments or projects; organisation-wide adoption absent.", 1),
        (f"D: No implementation of {topic}.", 0),
    ]

def _choices_measure(topic):
    return [
        (f"A: Quantitative KPIs for {topic} tracked systematically with defined methodology, baseline, and targets.", 5),
        (f"B: Periodic measurement of {topic} with limited KPIs; data collection inconsistent.", 3),
        (f"C: Minimal tracking of {topic}; no formal KPIs or documented methodology.", 1),
        (f"D: No measurement of {topic}.", 0),
    ]

def _choices_results(topic):
    return [
        (f"A: Results for {topic} disclosed publicly, show an improving trend, and are verified by third parties.", 5),
        (f"B: Results for {topic} tracked internally with limited public disclosure; trend partially visible.", 3),
        (f"C: Limited results data for {topic}; no clear performance trend or public disclosure.", 1),
        (f"D: No results data available for {topic}.", 0),
    ]

LAYER_NAMES = ["Policy", "Implementation", "Measurement", "Results"]
LAYER_FUNCS = [_choices_policy, _choices_impl, _choices_measure, _choices_results]

# ── Question definitions ──────────────────────────────────────────────────────

GRI1_CRITERIA = [
    ("GRI1-1", "Mandatory Requirements Application",
     "application of GRI Standards mandatory requirements",
     "implementation of GRI mandatory requirements",
     "compliance with GRI mandatory requirements",
     "evidence of GRI Standards conformance"),
    ("GRI1-2", "Responsible GRI Claims",
     "authorised use of GRI claims and statements of use",
     "preparation and publication of the GRI Statement of Use",
     "accuracy of GRI-related claims in external communications",
     "published GRI Statement of Use aligned with GRI 1 requirements"),
    ("GRI1-3", "Due Diligence Integration",
     "embedding human rights and environmental due diligence",
     "due diligence processes across the value chain",
     "due diligence outcomes and identified impacts",
     "documented results of ongoing due diligence processes"),
    ("GRI1-4", "Accuracy & Balance",
     "ensuring accuracy and balance in sustainability disclosures",
     "data collection and validation procedures for accurate reporting",
     "data quality controls and error detection mechanisms",
     "demonstrated accuracy and balanced presentation in published reports"),
    ("GRI1-5", "Completeness of Reporting",
     "defining reporting scope, boundary, and time period",
     "capturing complete data across all entities and value-chain tiers",
     "gap analysis and completeness assessment of reported information",
     "comprehensive coverage of all material impacts in published disclosures"),
    ("GRI1-6", "Comparability & Consistency",
     "consistent reporting methods across periods and entities",
     "standardised data-collection protocols enabling year-on-year comparison",
     "year-on-year comparisons and consistency checks",
     "comparative data published with restatements disclosed"),
    ("GRI1-7", "Sustainability Context",
     "contextualising performance against planetary boundaries and sector benchmarks",
     "integration of macro sustainability trends into reporting narrative",
     "monitoring performance relative to global and sector sustainability targets",
     "contextualised performance data published with reference to external benchmarks"),
    ("GRI1-8", "Stakeholder Inclusiveness",
     "identification of stakeholders and their legitimate interests",
     "engagement processes to incorporate stakeholder views",
     "tracking stakeholder feedback and its influence on reporting decisions",
     "evidence of stakeholder-responsive reporting with documented outcomes"),
]

GRI2_CRITERIA = [
    ("GRI2-1", "Organisational Profile",
     "disclosing organisational name, legal form, ownership, and location of headquarters",
     "maintaining accurate and current organisational profile information",
     "verification of organisational profile data accuracy",
     "complete, published organisational profile per GRI 2-1"),
    ("GRI2-2", "Entities Included in Report",
     "defining the reporting boundary (entities included and excluded)",
     "consistently applying the reporting boundary across all disclosures",
     "monitoring for boundary changes or restatements",
     "published list of all entities included in the sustainability report"),
    ("GRI2-3", "Reporting Period & Contact",
     "defining a consistent reporting period and designating a contact person",
     "maintaining up-to-date reporting period and contact information",
     "ensuring timely publication within the defined reporting period",
     "publicly available reporting period and contact details per GRI 2-3"),
    ("GRI2-4", "Restatements of Information",
     "documenting and disclosing restatements of previously reported data",
     "processes for identifying and correcting historical data errors",
     "tracking restatements and their materiality",
     "disclosed restatements with explanations in the sustainability report"),
    ("GRI2-5", "External Assurance",
     "commissioning external assurance of sustainability disclosures",
     "engaging qualified assurance providers and managing the assurance process",
     "monitoring assurance coverage (reasonable vs. limited) and findings",
     "publicly available assurance statement with scope, conclusions, and provider"),
    ("GRI2-6", "Activities & Value Chain",
     "mapping the organisation's activities, products/services, and value-chain relationships",
     "documenting value-chain partners and business relationships",
     "monitoring significant changes in the value chain",
     "published, comprehensive description of activities and value chain per GRI 2-6"),
    ("GRI2-7", "Employee Data",
     "tracking total employees by employment type, contract, and region",
     "maintaining current employee headcount and demographic data",
     "monitoring employee turnover, contract types, and regional distribution",
     "published employee data with breakdowns per GRI 2-7"),
    ("GRI2-8", "Non-Employee Workers",
     "identifying and tracking workers not employed directly (contractors, agency staff)",
     "integrating non-employee worker data into workforce reporting",
     "monitoring health, safety, and labour standards for non-employees",
     "published data on non-employee workers per GRI 2-8"),
    ("GRI2-9", "Governance Structure & Composition",
     "documenting the highest governance body structure and committee composition",
     "maintaining accurate governance structure records",
     "monitoring governance structure for independence and diversity",
     "published governance structure and composition per GRI 2-9"),
    ("GRI2-10", "Nomination & Selection of Governance Body",
     "nominating and selecting members of the highest governance body",
     "applying nomination criteria for governance independence and expertise",
     "evaluating governance body composition against best-practice criteria",
     "published nomination and selection process and outcome per GRI 2-10"),
    ("GRI2-12", "Board Role in Impact Oversight",
     "assigning governance oversight responsibility for material impacts",
     "ensuring the highest body reviews material impact disclosures",
     "tracking board engagement with impact topics and decisions made",
     "documented evidence of board oversight of impacts per GRI 2-12"),
    ("GRI2-13", "Delegation of Responsibility",
     "delegating management-level responsibility for impact management",
     "establishing clear accountability for sustainability across management",
     "monitoring whether delegated responsibilities are exercised effectively",
     "published accountability structure for sustainability per GRI 2-13"),
    ("GRI2-14", "Board Role in Sustainability Reporting",
     "assigning board responsibility for reviewing sustainability report content",
     "board review and approval process for the sustainability report",
     "tracking board input and corrections to sustainability disclosures",
     "documented board approval of sustainability report per GRI 2-14"),
    ("GRI2-16", "Critical Concerns to Highest Body",
     "establishing channels for communicating critical concerns to the board",
     "operating escalation processes that bring critical concerns to board attention",
     "tracking the number and nature of critical concerns escalated",
     "published number and description of critical concerns communicated per GRI 2-16"),
    ("GRI2-17", "Knowledge of Highest Governance Body",
     "ensuring board members have access to sustainability knowledge and training",
     "providing training and briefings on sustainability topics to board members",
     "assessing board knowledge levels on sustainability and ESG topics",
     "evidence of board competence in sustainability oversight per GRI 2-17"),
    ("GRI2-18", "Evaluation of Governance Performance",
     "conducting regular governance body performance evaluations",
     "implementing structured evaluation of the board on sustainability oversight",
     "tracking findings and improvement actions from governance evaluations",
     "published results of governance performance evaluation per GRI 2-18"),
    ("GRI2-22", "Statement on Sustainable Development Strategy",
     "publishing a CEO/senior leader statement linking strategy to sustainable development",
     "integrating sustainable development into the corporate strategy",
     "monitoring strategic targets aligned with sustainable development",
     "published CEO statement and strategy with measurable SD commitments per GRI 2-22"),
    ("GRI2-23", "Policy Commitments on Business Conduct",
     "establishing policy commitments on human rights, anti-corruption, and responsible business conduct",
     "communicating and embedding policy commitments throughout the value chain",
     "monitoring compliance with policy commitments and identifying deviations",
     "publicly available policy commitments with implementation evidence per GRI 2-23"),
    ("GRI2-29", "Approach to Stakeholder Engagement",
     "identifying stakeholder groups and their interests",
     "operating structured stakeholder engagement processes",
     "tracking stakeholder engagement frequency, methods, and topics raised",
     "published stakeholder engagement approach and outcomes per GRI 2-29"),
    ("GRI2-30", "Collective Bargaining Agreements",
     "recognising employee rights to collective bargaining",
     "negotiating and maintaining collective bargaining agreements",
     "monitoring percentage of employees covered by collective agreements",
     "published data on collective bargaining coverage per GRI 2-30"),
]

GRI3_CRITERIA = [
    ("GRI3-1a", "Impact Identification — Mapping Activities & Relationships",
     "mapping value-chain activities and relationships for impact identification",
     "conducting impact identification exercises across own operations and value chain",
     "tracking identified actual and potential impacts by type and severity",
     "published list of identified impacts with methodology per GRI 3-1"),
    ("GRI3-1b", "Impact Assessment — Severity & Likelihood",
     "assessing the severity and likelihood of identified impacts",
     "applying a documented impact-assessment methodology",
     "monitoring assessment results and updating impact rankings periodically",
     "published impact severity and likelihood assessments per GRI 3-1"),
    ("GRI3-1c", "Stakeholder Consultation in Materiality",
     "consulting affected stakeholders and experts in the materiality process",
     "conducting formal stakeholder consultation for materiality determination",
     "tracking stakeholder input and its influence on material topic selection",
     "published evidence of stakeholder consultation in materiality per GRI 3-1"),
    ("GRI3-2", "List of Material Topics",
     "determining and documenting the list of material topics",
     "maintaining and updating the material topic list after each cycle",
     "reviewing and validating the material topic list with senior leadership",
     "publicly disclosed, updated list of material topics per GRI 3-2"),
    ("GRI3-3a", "Management Approach — Policy",
     "adopting a management approach policy for each material topic",
     "communicating management approach policies internally and externally",
     "monitoring whether management approach policies are current and adequate",
     "published management approach policies for all material topics per GRI 3-3a"),
    ("GRI3-3b", "Management Approach — Topic Boundary",
     "defining the boundary (where impacts occur) for each material topic",
     "applying consistent boundary definitions in topic disclosures",
     "monitoring changes in topic boundaries over reporting periods",
     "published topic boundary definitions for all material topics per GRI 3-3b"),
    ("GRI3-3c", "Specific Actions to Address Impacts",
     "developing action plans to prevent, mitigate, or remediate material impacts",
     "implementing specific actions and programmes for each material topic",
     "tracking action implementation progress against defined timelines",
     "published specific actions and results for all material topics per GRI 3-3c"),
    ("GRI3-3d", "Tracking Effectiveness of Actions",
     "defining KPIs to measure effectiveness of management approach actions",
     "collecting data and reporting on effectiveness KPIs",
     "reviewing KPI results and adjusting actions based on findings",
     "published effectiveness KPIs with trend data per GRI 3-3d"),
    ("GRI3-3e", "Remediation Mechanisms",
     "establishing remediation mechanisms for negative impacts",
     "operating accessible and fair remediation processes",
     "monitoring remediation cases: number, nature, and resolution status",
     "published number and outcomes of remediation cases per GRI 3-3e"),
    ("GRI3-3f", "Grievance Mechanisms",
     "establishing grievance mechanisms for affected stakeholders",
     "publicising and operating grievance mechanisms accessible to all affected parties",
     "tracking grievances received, processed, and resolved",
     "published grievance mechanism results and access information per GRI 3-3f"),
    ("GRI3-3g", "Engagement with Affected Stakeholders on Topics",
     "engaging affected stakeholders in the management of each material topic",
     "incorporating affected stakeholder feedback into topic management decisions",
     "tracking engagement quality and stakeholder satisfaction",
     "evidence of ongoing stakeholder engagement influencing topic management per GRI 3-3g"),
    ("GRI3-3h", "Lessons Learned & Improvement",
     "capturing lessons learned from impact incidents and near-misses",
     "integrating lessons learned into management approach revisions",
     "tracking number of improvement actions driven by lessons learned",
     "documented improvement cycle linked to lessons learned per GRI 3-3h"),
    ("GRI3-3i", "Supply Chain Material Impacts",
     "assessing material impacts within the supply chain",
     "implementing supplier requirements linked to material topics",
     "monitoring supplier compliance with material topic requirements",
     "published supply chain impact data and supplier performance per GRI 3-3i"),
    ("GRI3-3j", "Human Rights Due Diligence in Material Topics",
     "conducting human rights due diligence within material topic management",
     "implementing human rights safeguards in operations and supply chain",
     "tracking human rights incidents and mitigation effectiveness",
     "published human rights performance data linked to material topics per GRI 3-3j"),
    ("GRI3-gri-index", "GRI Content Index Completeness",
     "committing to a complete and accurate GRI Content Index",
     "compiling and cross-referencing all required GRI disclosures in the index",
     "reviewing GRI Content Index for accuracy and omissions",
     "published, complete GRI Content Index with all required disclosures per GRI 3"),
]

SECTOR_CRITERIA = {
    "Sector — Agriculture & Food": [
        ("AGRI-1", "Land Rights & Tenure"),
        ("AGRI-2", "Smallholder & Community Inclusion"),
        ("AGRI-3", "Water Use in Agriculture"),
        ("AGRI-4", "Biodiversity & Ecosystem Services"),
        ("AGRI-5", "Pesticide & Chemical Management"),
        ("AGRI-6", "Food Safety & Quality"),
        ("AGRI-7", "Child & Forced Labour in Supply Chain"),
        ("AGRI-8", "Greenhouse Gas Emissions — Agriculture"),
    ],
    "Sector — Energy & Utilities": [
        ("ENER-1", "GHG Emissions from Energy Generation"),
        ("ENER-2", "Renewable Energy Transition"),
        ("ENER-3", "Water Consumption in Energy Processes"),
        ("ENER-4", "Physical Climate Risk Management"),
        ("ENER-5", "Just Transition for Workers"),
        ("ENER-6", "Energy Access & Affordability"),
        ("ENER-7", "Air Quality & Pollutant Emissions"),
        ("ENER-8", "Nuclear Safety & Radioactive Waste"),
    ],
    "Sector — Financial Services": [
        ("FIN-1", "Financed Emissions (PCAF)"),
        ("FIN-2", "Climate Risk in Lending Portfolio"),
        ("FIN-3", "Responsible Investment Policy"),
        ("FIN-4", "Financial Inclusion & Access"),
        ("FIN-5", "Anti-Money Laundering & Compliance"),
        ("FIN-6", "Data Privacy & Cybersecurity"),
        ("FIN-7", "Diversity in Financial Services"),
        ("FIN-8", "ESG Integration in Products"),
    ],
    "Sector — Manufacturing & Indust": [
        ("MFG-1", "Energy Efficiency in Manufacturing"),
        ("MFG-2", "Waste Generation & Circularity"),
        ("MFG-3", "Chemical & Hazardous Materials Safety"),
        ("MFG-4", "Occupational Health & Safety"),
        ("MFG-5", "Supplier Labour Standards"),
        ("MFG-6", "Product Stewardship & Lifecycle"),
        ("MFG-7", "GHG Emissions — Manufacturing Processes"),
        ("MFG-8", "Community Impact of Operations"),
    ],
    "Sector — Construction & Real Es": [
        ("CON-1", "Green Building Standards Adoption"),
        ("CON-2", "Embodied Carbon in Materials"),
        ("CON-3", "Biodiversity Impact of Development"),
        ("CON-4", "Waste & Demolition Materials"),
        ("CON-5", "Worker Safety on Construction Sites"),
        ("CON-6", "Affordable Housing & Community Need"),
        ("CON-7", "Climate Resilience of Built Assets"),
        ("CON-8", "Land & Resource Rights"),
    ],
    "Sector — Healthcare & Pharma": [
        ("HLT-1", "Access to Medicines & Pricing"),
        ("HLT-2", "Clinical Trial Transparency"),
        ("HLT-3", "Antimicrobial Resistance Management"),
        ("HLT-4", "Medical Waste & Hazardous Disposal"),
        ("HLT-5", "Patient Privacy & Data Protection"),
        ("HLT-6", "Health Worker Safety"),
        ("HLT-7", "Product Quality & Safety"),
        ("HLT-8", "Ethical Marketing & Promotion"),
    ],
    "Sector — Technology & IT": [
        ("TECH-1", "Data Privacy & Security"),
        ("TECH-2", "Algorithmic Accountability"),
        ("TECH-3", "E-Waste & Product Take-Back"),
        ("TECH-4", "Energy Use of Data Centres"),
        ("TECH-5", "Digital Inclusion & Access"),
        ("TECH-6", "Responsible AI & Automation"),
        ("TECH-7", "Conflict Minerals in Supply Chain"),
        ("TECH-8", "Cybersecurity Risk Management"),
    ],
    "Sector — Retail & Trade": [
        ("RET-1", "Supply Chain Labour Standards"),
        ("RET-2", "Product Safety & Recalls"),
        ("RET-3", "Packaging & Plastic Waste"),
        ("RET-4", "Consumer Privacy & Data"),
        ("RET-5", "Responsible Sourcing of Raw Materials"),
        ("RET-6", "Food Waste in Retail Operations"),
        ("RET-7", "Living Wage in Supply Chain"),
        ("RET-8", "Carbon Footprint of Logistics"),
    ],
}


# ── Excel builder helpers ──────────────────────────────────────────────────────

HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF")
SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
SUBHDR_FONT = Font(bold=True)

def _write_header(ws, row=1):
    ws.cell(row=row, column=1, value="#")
    ws.cell(row=row, column=2, value="ID")
    ws.cell(row=row, column=3, value="Category")
    ws.cell(row=row, column=4, value="Layer")
    ws.cell(row=row, column=5, value="Criterion Description")
    ws.cell(row=row, column=6, value="Question Text")
    ws.cell(row=row, column=7, value="Ref")
    ws.cell(row=row, column=8, value="Option")
    ws.cell(row=row, column=9, value="Score")
    for c in range(1, 10):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["H"].width = 80
    ws.column_dimensions["I"].width = 8


def _write_core_criteria(ws, criteria_list, start_row=2):
    """Write P/I/M/R questions for core sheets."""
    row = start_row
    q_texts = [
        "What is the organisation's policy for {topic}?",
        "How is the organisation implementing its approach to {topic}?",
        "How does the organisation measure progress on {topic}?",
        "What results has the organisation achieved on {topic}?",
    ]
    for qid, name, p_topic, i_topic, m_topic, r_topic in criteria_list:
        topics = [p_topic, i_topic, m_topic, r_topic]
        for layer_idx, (layer_name, fn) in enumerate(zip(LAYER_NAMES, LAYER_FUNCS)):
            topic = topics[layer_idx]
            choices = fn(topic)
            # Unique sub-question ID
            layer_code = layer_name[0]  # P / I / M / R
            sub_qid = f"{qid}-{layer_code}"
            q_text = q_texts[layer_idx].format(topic=topic)
            for choice_idx, (choice_text, score) in enumerate(choices):
                ws.cell(row=row, column=1, value=row - start_row + 1)
                ws.cell(row=row, column=2, value=sub_qid)
                ws.cell(row=row, column=3, value=name)
                ws.cell(row=row, column=4, value=layer_name)
                ws.cell(row=row, column=5, value=f"{qid}: {name}")
                ws.cell(row=row, column=6, value=q_text)
                ws.cell(row=row, column=7, value=qid)
                ws.cell(row=row, column=8, value=choice_text)
                ws.cell(row=row, column=9, value=score)
                row += 1
    return row


def _write_sector_criteria(ws, criteria_list, start_row=2):
    """Write sector sheets — one simple choice question per criterion, 4 options."""
    row = start_row
    sector_choices = [
        ("A: Formal policy with measurable targets, implemented across all operations, with publicly reported results.", 5),
        ("B: Policy exists and partially implemented; limited measurement and inconsistent external reporting.", 3),
        ("C: Informal approach to this topic; minimal measurement and no public disclosure.", 1),
        ("D: No approach, policy, or reporting on this topic.", 0),
    ]
    for qid, name in criteria_list:
        q_text = f"How does the organisation manage and report on {name}?"
        for choice_text, score in sector_choices:
            ws.cell(row=row, column=1, value=row - start_row + 1)
            ws.cell(row=row, column=2, value=qid)
            ws.cell(row=row, column=3, value=name)
            ws.cell(row=row, column=4, value="Sector")
            ws.cell(row=row, column=5, value=name)
            ws.cell(row=row, column=6, value=q_text)
            ws.cell(row=row, column=7, value=qid)
            ws.cell(row=row, column=8, value=choice_text)
            ws.cell(row=row, column=9, value=score)
            row += 1
    return row


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── GRI 1 —Foundation ────────────────────────────────────────────────────
    ws1 = wb.create_sheet("GRI 1 — Foundation")
    _write_header(ws1)
    r = _write_core_criteria(ws1, GRI1_CRITERIA)
    print(f"  GRI 1 — Foundation: {r - 2} rows  ({len(GRI1_CRITERIA)*4} questions × 4 choices)")

    # ── GRI 2 — General Disclosures ──────────────────────────────────────────
    ws2 = wb.create_sheet("GRI 2 — General Disclosures")
    _write_header(ws2)
    r = _write_core_criteria(ws2, GRI2_CRITERIA)
    print(f"  GRI 2 — General Disclosures: {r - 2} rows  ({len(GRI2_CRITERIA)*4} questions × 4 choices)")

    # ── GRI 3 — Material Topics ──────────────────────────────────────────────
    ws3 = wb.create_sheet("GRI 3 — Material Topics")
    _write_header(ws3)
    r = _write_core_criteria(ws3, GRI3_CRITERIA)
    print(f"  GRI 3 — Material Topics: {r - 2} rows  ({len(GRI3_CRITERIA)*4} questions × 4 choices)")

    # ── Sector sheets ─────────────────────────────────────────────────────────
    for sheet_name, criteria in SECTOR_CRITERIA.items():
        ws = wb.create_sheet(sheet_name)
        # Sector sheets need "Q ID" in column B for _parse_sector()
        ws.cell(row=1, column=1, value="#")
        ws.cell(row=1, column=2, value="Q ID")
        ws.cell(row=1, column=3, value="Category")
        ws.cell(row=1, column=4, value="Layer")
        ws.cell(row=1, column=5, value="Criterion")
        ws.cell(row=1, column=6, value="Option")
        ws.cell(row=1, column=7, value="Score")
        for c in range(1, 8):
            cell = ws.cell(row=1, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["F"].width = 80
        ws.column_dimensions["G"].width = 8

        # _parse_sector expects: p[1]=q_id, p[2]=cat_name, p[5]=option, p[6]=pts
        row = 2
        sector_choices = [
            ("A: Formal policy with measurable targets, fully implemented and publicly reported.", 5),
            ("B: Policy exists and partially implemented; limited measurement or incomplete reporting.", 3),
            ("C: Informal approach; minimal measurement and no systematic public disclosure.", 1),
            ("D: No policy, implementation, or reporting on this topic.", 0),
        ]
        for qid, name in criteria:
            for choice_text, score in sector_choices:
                ws.cell(row=row, column=1, value=row - 1)
                ws.cell(row=row, column=2, value=qid)
                ws.cell(row=row, column=3, value=name)
                ws.cell(row=row, column=4, value="Sector")
                ws.cell(row=row, column=5, value=name)
                ws.cell(row=row, column=6, value=choice_text)
                ws.cell(row=row, column=7, value=score)
                row += 1
        print(f"  {sheet_name}: {row - 2} rows ({len(criteria)} questions × 4 choices)")

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\n✅  Saved: {OUT_PATH}")
    print(f"   Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
